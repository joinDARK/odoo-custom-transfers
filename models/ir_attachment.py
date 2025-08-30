# -*- coding: utf-8 -*-
import io
import base64
import logging
import requests
from io import BytesIO
from collections import defaultdict

import pandas as pd
from docx import Document as DocxDocument
from odoo import api, fields, models, _
from odoo.exceptions import AccessError

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º logger –ü–ï–†–í–´–ú!
_logger = logging.getLogger(__name__)

# –î–ò–ê–ì–ù–û–°–¢–ò–ß–ï–°–ö–û–ï –°–û–û–ë–©–ï–ù–ò–ï - –ü–†–û–í–ï–†–Ø–ï–ú –ó–ê–ì–†–£–ñ–ê–ï–¢–°–Ø –õ–ò –§–ê–ô–õ
_logger.info("üö® AMANAT: ir_attachment.py module is being loaded!")
print("üö® AMANAT: ir_attachment.py module is being loaded!")  # –î—É–±–ª–∏—Ä—É–µ–º –≤ print

# –ü–æ–ø—ã—Ç–∫–∞ –∏–º–ø–æ—Ä—Ç–∞ —Ä–∞–∑–ª–∏—á–Ω—ã—Ö –¥–≤–∏–∂–∫–æ–≤ –¥–ª—è —á—Ç–µ–Ω–∏—è Excel
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


class IrAttachment(models.Model):
    """Extended Attachment Model for Amanat Sverka Files"""
    _inherit = 'ir.attachment'

    def _check_access(self, operation: str):
        """
        üö®üö®üö® –°–£–ü–ï–†-–Ø–î–ï–†–ù–´–ô –û–ë–•–û–î: –í—Å–µ–≥–¥–∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º None = –Ω–µ—Ç –ø—Ä–æ–±–ª–µ–º —Å –¥–æ—Å—Ç—É–ø–æ–º üö®üö®üö®
        """
        _logger.error(f"üö®üö®üö® AMANAT ir.attachment._check_access CALLED! operation: {operation}, user: {self.env.user.name}, IDs: {self.ids} üö®üö®üö®")
        _logger.error(f"üö® AMANAT: ALWAYS RETURNING None - FULL ACCESS TO ALL ATTACHMENTS! üö®")
        print(f"üö® AMANAT: ir.attachment._check_access NUCLEAR BYPASS for {self.env.user.name}, operation: {operation} üö®")
        
        # –°–£–ü–ï–†-–Ø–î–ï–†–ù–´–ô –ü–û–î–•–û–î: –í—Å–µ–≥–¥–∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º None = –Ω–µ—Ç –ø—Ä–æ–±–ª–µ–º —Å –¥–æ—Å—Ç—É–ø–æ–º
        return None

    


    
    def _filter_access_rules(self, operation):
        """–Ø–î–ï–†–ù–´–ô –û–ë–•–û–î: –ü–æ–ª–Ω–æ—Å—Ç—å—é –æ—Ç–∫–ª—é—á–∞–µ–º record rules –¥–ª—è –í–°–ï–• –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
        _logger.error(f"üö®üö®üö® AMANAT _filter_access_rules called for IDs: {self.ids}, operation: {operation}, user: {self.env.user.name} üö®üö®üö®")
        
        # –≠–ö–°–¢–†–ï–ú–ê–õ–¨–ù–û–ï –ò–ó–ú–ï–ù–ï–ù–ò–ï: –û–±—Ö–æ–¥–∏–º –ø—Ä–∞–≤–∏–ª–∞ –¥–ª—è –í–°–ï–• –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –±–µ–∑ –∏—Å–∫–ª—é—á–µ–Ω–∏–π
        _logger.error(f"üö® AMANAT: BYPASSING ALL RECORD RULES FOR ALL USERS - NO EXCEPTIONS! üö®")
        print(f"üö® AMANAT: _filter_access_rules BYPASS for user {self.env.user.name}, operation: {operation} üö®")
        
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –≤—Å–µ –∑–∞–ø–∏—Å–∏ –±–µ–∑ —Ñ–∏–ª—å—Ç—Ä–∞—Ü–∏–∏ –¥–ª—è –í–°–ï–•
        return self

    def read(self, fields=None, load='_classic_read'):
        """
        AMANAT: –ê–≥—Ä–µ—Å—Å–∏–≤–Ω–æ–µ –ø–µ—Ä–µ–æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏–µ read() –¥–ª—è –æ–±—Ö–æ–¥–∞ record rules
        """
        _logger.info("AMANAT ir.attachment.read() called for IDs: %s, user: %s", self.ids, self.env.user.name)
        
        if self.env.user._is_superuser():
            _logger.info("AMANAT: Superuser - using normal read() for %s", self.ids)
            return super().read(fields, load)
            
        if self.env.user.has_group('base.group_user'):
            _logger.info("AMANAT: Internal user %s - BYPASSING RECORD RULES with sudo() for %s", self.env.user.name, self.ids)
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º sudo() –¥–ª—è –æ–±—Ö–æ–¥–∞ record rules, –Ω–æ –≤—ã–∑—ã–≤–∞–µ–º –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω—ã–π –º–µ—Ç–æ–¥
            return super(IrAttachment, self.sudo()).read(fields, load)
            
        _logger.info("AMANAT: External user %s - using normal read() for %s", self.env.user.name, self.ids)
        return super().read(fields, load)

    @api.model
    def check(self, mode, values=None):
        """
        üö®üö®üö® –≠–ö–°–¢–†–ï–ú–ê–õ–¨–ù–´–ô –Ø–î–ï–†–ù–´–ô –û–ë–•–û–î: –ü–û–õ–ù–´–ô –î–û–°–¢–£–ü –ö–û –í–°–ï–ú –§–ê–ô–õ–ê–ú üö®üö®üö®
        –≠–¢–û–¢ –ú–ï–¢–û–î –ü–ï–†–ï–û–ü–†–ï–î–ï–õ–Ø–ï–¢ –ë–ê–ó–û–í–´–ô ir.attachment.check() –ò–ó ODOO CORE
        """
        _logger.error(f"üö®üö®üö® AMANAT OVERRIDDEN CHECK() CALLED! IDs: {self.ids}, mode: {mode}, user: {self.env.user.name} üö®üö®üö®")
        print(f"üö®üö®üö® AMANAT OVERRIDDEN CHECK() CALLED! IDs: {self.ids}, mode: {mode}, user: {self.env.user.name} üö®üö®üö®")
        
        # –ê–ë–°–û–õ–Æ–¢–ù–û –Ø–î–ï–†–ù–´–ô –ü–û–î–•–û–î: –†–∞–∑—Ä–µ—à–∞–µ–º –í–°–ï –æ–ø–µ—Ä–∞—Ü–∏–∏ –¥–ª—è –í–°–ï–• –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
        _logger.error(f"üö® AMANAT: FULL ACCESS GRANTED TO ALL USERS - BYPASSING ALL SECURITY! üö®")
        print(f"üö® AMANAT: FULL ACCESS GRANTED TO ALL USERS - BYPASSING ALL SECURITY! üö®")
        return True
        

    @api.model
    def get_hidden_columns_and_data_range(self, xlsx_data):
        """–û–ø—Ä–µ–¥–µ–ª—è–µ—Ç —Å–∫—Ä—ã—Ç—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –∏ —Ä–µ–∞–ª—å–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏ —á—Ç–µ–Ω–∏—è"""
        if not OPENPYXL_AVAILABLE:
            return set(), None, None
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            # –ó–∞–≥—Ä—É–∂–∞–µ–º —Ñ–∞–π–ª —Ç–æ–ª—å–∫–æ –¥–ª—è –∞–Ω–∞–ª–∏–∑–∞ —Å—Ç—Ä—É–∫—Ç—É—Ä—ã
            workbook = openpyxl.load_workbook(
                BytesIO(xlsx_data), 
                data_only=True,
                read_only=True
            )
            
            sheet = workbook.active
            if not sheet:
                sheet = workbook[workbook.sheetnames[0]]
            
            hidden_columns = set()
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ä–µ–∞–ª—å–Ω—ã–µ –≥—Ä–∞–Ω–∏—Ü—ã —Ñ–∞–π–ª–∞
            # –ò—Å–ø–æ–ª—å–∑—É–µ–º max_column –∏–∑ –æ–øenpyxl –∫–∞–∫ –º–∞–∫—Å–∏–º–∞–ª—å–Ω—É—é –≥—Ä–∞–Ω–∏—Ü—É
            real_max_column = sheet.max_column or 1
            real_max_row = sheet.max_row or 1
            
            # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –∞–Ω–∞–ª–∏–∑ —Ä–∞–∑—É–º–Ω—ã–º–∏ –ø—Ä–µ–¥–µ–ª–∞–º–∏
            max_cols_to_analyze = min(real_max_column, 50)  # –ú–∞–∫—Å–∏–º—É–º 50 –∫–æ–ª–æ–Ω–æ–∫
            max_rows_to_analyze = min(real_max_row, 100)    # –ú–∞–∫—Å–∏–º—É–º 100 —Å—Ç—Ä–æ–∫ –¥–ª—è –∞–Ω–∞–ª–∏–∑–∞
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–∏–π –¥–∏–∞–ø–∞–∑–æ–Ω –¥–∞–Ω–Ω—ã—Ö
            actual_data_column = 0
            for row in sheet.iter_rows(min_row=1, max_row=max_rows_to_analyze, max_col=max_cols_to_analyze, values_only=True):
                for col_idx, cell_value in enumerate(row):
                    if cell_value is not None and str(cell_value).strip():
                        actual_data_column = max(actual_data_column, col_idx + 1)
            
            # –ï—Å–ª–∏ –¥–∞–Ω–Ω—ã—Ö –Ω–µ –Ω–∞–π–¥–µ–Ω–æ, –∏—Å–ø–æ–ª—å–∑—É–µ–º –º–∏–Ω–∏–º–∞–ª—å–Ω—ã–π –¥–∏–∞–ø–∞–∑–æ–Ω
            if actual_data_column == 0:
                actual_data_column = min(real_max_column, 10)  # –ú–∏–Ω–∏–º—É–º 10 –∫–æ–ª–æ–Ω–æ–∫
            
            # –£–±–µ–∂–¥–∞–µ–º—Å—è, —á—Ç–æ actual_data_column –Ω–µ –ø—Ä–µ–≤—ã—à–∞–µ—Ç —Ä–µ–∞–ª—å–Ω—ã–π –º–∞–∫—Å–∏–º—É–º
            actual_data_column = min(actual_data_column, real_max_column)
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Å–∫—Ä—ã—Ç—ã–µ –∫–æ–ª–æ–Ω–∫–∏ —Ç–æ–ª—å–∫–æ –≤ –ø—Ä–µ–¥–µ–ª–∞—Ö —Ñ–∞–∫—Ç–∏—á–µ—Å–∫–æ–≥–æ –¥–∏–∞–ø–∞–∑–æ–Ω–∞
            for col_num in range(1, actual_data_column + 1):
                column_letter = get_column_letter(col_num)
                
                # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–∞–∑–ª–∏—á–Ω—ã–µ —Å–ø–æ—Å–æ–±—ã —Å–∫—Ä—ã—Ç–∏—è –∫–æ–ª–æ–Ω–æ–∫
                is_hidden = False
                
                # –°–ø–æ—Å–æ–± 1: —á–µ—Ä–µ–∑ column_dimensions.hidden
                if column_letter in sheet.column_dimensions:
                    if getattr(sheet.column_dimensions[column_letter], 'hidden', False):
                        is_hidden = True
                
                # –°–ø–æ—Å–æ–± 2: —á–µ—Ä–µ–∑ width = 0 (–∏–Ω–æ–≥–¥–∞ —Ç–∞–∫ —Å–∫—Ä—ã–≤–∞—é—Ç –∫–æ–ª–æ–Ω–∫–∏)
                if column_letter in sheet.column_dimensions:
                    width = getattr(sheet.column_dimensions[column_letter], 'width', None)
                    if width is not None and width == 0:
                        is_hidden = True
                
                # –°–ø–æ—Å–æ–± 3: –ø—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –∫–æ–ª–æ–Ω–∫–∞ –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–æ –ø—É—Å—Ç–∞—è
                if not is_hidden:
                    col_values = []
                    for row_num in range(1, min(21, max_rows_to_analyze + 1)):
                        if row_num <= real_max_row:  # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≥—Ä–∞–Ω–∏—Ü—ã
                            cell_value = sheet.cell(row=row_num, column=col_num).value
                            if cell_value is not None and str(cell_value).strip():
                                col_values.append(str(cell_value).strip())
                    
                    # –ï—Å–ª–∏ –∫–æ–ª–æ–Ω–∫–∞ –ø–æ–ª–Ω–æ—Å—Ç—å—é –ø—É—Å—Ç–∞—è, —Å—á–∏—Ç–∞–µ–º –µ—ë —Å–∫—Ä—ã—Ç–æ–π
                    if len(col_values) == 0:
                        is_hidden = True
                
                if is_hidden:
                    hidden_columns.add(col_num - 1)  # –ò–Ω–¥–µ–∫—Å —Å 0
            
            workbook.close()
            
            return hidden_columns, actual_data_column, min(real_max_row, 5000)
            
        except Exception as e:
            # –ï—Å–ª–∏ –Ω–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ–∞–Ω–∞–ª–∏–∑–∏—Ä–æ–≤–∞—Ç—å —Å—Ç—Ä—É–∫—Ç—É—Ä—É, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –±–µ–∑–æ–ø–∞—Å–Ω—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è
            return set(), 30, 5000  # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º 30 –∫–æ–ª–æ–Ω–∫–∞–º–∏ –∏ 5000 —Å—Ç—Ä–æ–∫–∞–º–∏

    @api.model
    def decode_content(self, attach_id, doc_type):
        """Decode XLSX, XLS, or DOCX File Data.
        This method takes a binary file data from an attachment and decodes
        the content of the file for XLSX, XLS, and DOCX file formats.
        :param int attach_id: id of attachment.
        :param str doc_type: the type of the given attachment either 'xlsx', 'xls', or 'docx'
        :return: return the decoded data."""
        try:
            attachment = self.sudo().browse(attach_id)
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Ñ–∞–π–ª —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
            if not attachment.exists():
                return ("<p style='padding-top:8px;color:red;'>"
                        "–§–∞–π–ª –Ω–µ –Ω–∞–π–¥–µ–Ω</p>")
            
            # –ü–æ–ª—É—á–∞–µ–º –¥–∞–Ω–Ω—ã–µ —Ñ–∞–π–ª–∞
            xlsx_data = None
            
            # –°–ø–æ—Å–æ–± 1: –ü—Ä–æ–±—É–µ–º —Å–Ω–∞—á–∞–ª–∞ –ø–æ–ª—É—á–∏—Ç—å binary –¥–∞–Ω–Ω—ã–µ (–∫–∞–∫ chatter_attachment_manager)
            try:
                if attachment.datas:
                    xlsx_data = base64.b64decode(attachment.datas)
                    if xlsx_data:
                        # –£—Å–ø–µ—à–Ω–æ –ø–æ–ª—É—á–∏–ª–∏ binary –¥–∞–Ω–Ω—ã–µ
                        pass
            except Exception as e:
                pass
            
            # –°–ø–æ—Å–æ–± 2: –ï—Å–ª–∏ binary –¥–∞–Ω–Ω—ã—Ö –Ω–µ—Ç, –ø—Ä–æ–±—É–µ–º –∑–∞–≥—Ä—É–∑–∏—Ç—å –ø–æ URL
            if not xlsx_data and hasattr(attachment, 'type') and attachment.type == 'url':
                try:
                    if not attachment.url:
                        return ("<p style='padding-top:8px;color:red;'>"
                                "URL —Ñ–∞–π–ª–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω</p>")
                    
                    # –ó–∞–≥—Ä—É–∂–∞–µ–º —Ñ–∞–π–ª —Å –≤–Ω–µ—à–Ω–µ–≥–æ URL
                    response = requests.get(attachment.url, timeout=30)
                    response.raise_for_status()
                    xlsx_data = response.content
                    
                except requests.exceptions.RequestException as e:
                    return (f"<p style='padding-top:8px;color:red;'>"
                            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–≥—Ä—É–∑–∫–µ —Ñ–∞–π–ª–∞: {str(e)}</p>")
                except Exception as e:
                    return (f"<p style='padding-top:8px;color:red;'>"
                            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –¥–æ—Å—Ç—É–ø–µ –∫ URL: {str(e)}</p>")
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—É—á–µ–Ω—ã
            if not xlsx_data:
                return ("<p style='padding-top:8px;color:red;'>"
                        "–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —Ñ–∞–π–ª–∞</p>")
            
            # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–∞–π–ª—ã
            if doc_type in ['xlsx', 'xls', 'docx', 'pdf']:
                try:
                    content = pd.DataFrame()  # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º –ø–µ—Ä–µ–º–µ–Ω–Ω—É—é content
                    if doc_type == 'xlsx':
                        # –°–Ω–∞—á–∞–ª–∞ –∞–Ω–∞–ª–∏–∑–∏—Ä—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É —Ñ–∞–π–ª–∞ –¥–ª—è –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏
                        hidden_columns, max_data_column, max_data_row = self.get_hidden_columns_and_data_range(xlsx_data)
                        
                        # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —á–∏—Ç–∞–µ–º—ã—Ö –∫–æ–ª–æ–Ω–æ–∫ –∏ —Å—Ç—Ä–æ–∫
                        max_cols_to_read = min(max_data_column or 30, 50)  # –ú–∞–∫—Å–∏–º—É–º 50 –∫–æ–ª–æ–Ω–æ–∫
                        max_rows_to_read = min(max_data_row or 5000, 5000)  # –ú–∞–∫—Å–∏–º—É–º 5000 —Å—Ç—Ä–æ–∫
                        content = pd.DataFrame()  # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º –∫–∞–∫ –ø—É—Å—Ç–æ–π DataFrame
                        error_msgs = []
                        
                        # –°–ø–æ—Å–æ–± 1: openpyxl —Å –ø—Ä–∞–≤–∏–ª—å–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–æ–π —Ñ–æ—Ä–º—É–ª
                        if OPENPYXL_AVAILABLE:
                            try:
                                import openpyxl
                                from openpyxl.utils import get_column_letter
                                
                                # –ó–∞–≥—Ä—É–∂–∞–µ–º —Ñ–∞–π–ª —Å –ø–æ–ª–Ω–æ–π –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –æ —è—á–µ–π–∫–∞—Ö
                                workbook = openpyxl.load_workbook(
                                    BytesIO(xlsx_data), 
                                    data_only=False,     # –ü–æ–ª—É—á–∞–µ–º —Ñ–æ—Ä–º—É–ª—ã –∏ –∑–Ω–∞—á–µ–Ω–∏—è
                                    read_only=False      # –ù—É–∂–Ω–æ –¥–ª—è —Ä–∞–±–æ—Ç—ã —Å —Ñ–æ—Ä–º—É–ª–∞–º–∏
                                )
                                
                                sheet = workbook.active
                                if not sheet:
                                    sheet = workbook[workbook.sheetnames[0]]
                                
                                data = []
                                
                                # –ß–∏—Ç–∞–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ—Å—Ç—Ä–æ—á–Ω–æ —Å –ø—Ä–∞–≤–∏–ª—å–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–æ–π —Ñ–æ—Ä–º—É–ª
                                if sheet:
                                    for row_num, row_cells in enumerate(sheet.iter_rows(max_row=max_rows_to_read, max_col=max_cols_to_read)):
                                        if row_num > max_rows_to_read:
                                            break
                                        # –í–∫–ª—é—á–∞–µ–º –≤—Å–µ —Å—Ç—Ä–æ–∫–∏, –≤–∫–ª—é—á–∞—è –ø—É—Å—Ç—ã–µ (–∫–∞–∫ –≤ Excel)
                                        clean_row = []
                                        for cell_idx, cell in enumerate(row_cells):
                                            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–∫—Ä—ã—Ç—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                            if cell_idx in hidden_columns:
                                                continue
                                                
                                            try:
                                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–∏–ø —è—á–µ–π–∫–∏ –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–æ—Ä–º—É–ª—ã
                                                if cell.value is None:
                                                    clean_row.append("")
                                                elif isinstance(cell.value, str) and cell.value.startswith('='):
                                                    # –≠—Ç–æ —Ñ–æ—Ä–º—É–ª–∞ - –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –µ—ë –∫–∞–∫ –µ—Å—Ç—å
                                                    clean_row.append(str(cell.value))
                                                else:
                                                    # –û–±—ã—á–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
                                                    cell_value = str(cell.value)
                                                    # –£–±–∏—Ä–∞–µ–º –∞—Ä—Ç–µ—Ñ–∞–∫—Ç—ã pandas
                                                    if cell_value in ['nan', 'NaN', 'NaT', 'None']:
                                                        clean_row.append("")
                                                    else:
                                                        clean_row.append(cell_value)
                                            except:
                                                clean_row.append("")
                                        
                                        # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç—Ä–æ–∫–∏ —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –æ–Ω–∏ –Ω–µ –ø—É—Å—Ç—ã–µ –∏–ª–∏ —ç—Ç–æ –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ (–∑–∞–≥–æ–ª–æ–≤–∫–∏)
                                        if clean_row and (row_num == 0 or any(cell.strip() for cell in clean_row if cell)):
                                            data.append(clean_row)
                                
                                workbook.close()
                                
                                if data:
                                    # –°–æ–∑–¥–∞–µ–º DataFrame —Å –æ—á–∏—â–µ–Ω–Ω—ã–º–∏ –¥–∞–Ω–Ω—ã–º–∏
                                    if len(data) > 1:
                                        # –ü–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ –∫–∞–∫ –∑–∞–≥–æ–ª–æ–≤–∫–∏
                                        headers = []
                                        for i, header in enumerate(data[0]):
                                            if header and str(header).strip():
                                                headers.append(str(header).strip())
                                            else:
                                                headers.append(f"Column_{i+1}")
                                        
                                        content = pd.DataFrame(data[1:], columns=headers)
                                    else:
                                        content = pd.DataFrame(data)
                                        
                            except Exception as e1:
                                error_msgs.append(f"openpyxl —É–ª—É—á—à–µ–Ω–Ω—ã–π: {str(e1)}")
                        
                        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –†–ï–ê–õ–¨–ù–û–ï –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–æ–ª–æ–Ω–æ–∫ –≤ —Ñ–∞–π–ª–µ –Ω–µ—Å–∫–æ–ª—å–∫–∏–º–∏ —Å–ø–æ—Å–æ–±–∞–º–∏
                        actual_columns_count = None
                        
                        # –°–ø–æ—Å–æ–± 1: –ü—Ä–æ–±—É–µ–º —á–µ—Ä–µ–∑ pandas –±–µ–∑ –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤
                        try:
                            test_df = pd.read_excel(BytesIO(xlsx_data), engine='openpyxl', nrows=1, header=None)
                            actual_columns_count = len(test_df.columns)
                        except:
                            pass
                        
                        # –°–ø–æ—Å–æ–± 2: –ï—Å–ª–∏ –ø–µ—Ä–≤—ã–π —Å–ø–æ—Å–æ–± –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª, –ø—Ä–æ–±—É–µ–º —á–µ—Ä–µ–∑ openpyxl –Ω–∞–ø—Ä—è–º—É—é
                        if actual_columns_count is None and OPENPYXL_AVAILABLE:
                            try:
                                import openpyxl
                                workbook = openpyxl.load_workbook(BytesIO(xlsx_data), data_only=True, read_only=True)
                                sheet = workbook.active
                                if not sheet:
                                    sheet = workbook[workbook.sheetnames[0]]
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ä–µ–∞–ª—å–Ω–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∫–æ–ª–æ–Ω–æ–∫ —Å –¥–∞–Ω–Ω—ã–º–∏
                                max_col_with_data = 0
                                for row in sheet.iter_rows(min_row=1, max_row=5, values_only=True):
                                    for col_idx, cell_value in enumerate(row):
                                        if cell_value is not None:
                                            max_col_with_data = max(max_col_with_data, col_idx + 1)
                                
                                actual_columns_count = max_col_with_data
                                workbook.close()
                            except:
                                pass
                        
                        # –°–ø–æ—Å–æ–± 3: –ï—Å–ª–∏ –≤—Å—ë –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª–æ, –∏—Å–ø–æ–ª—å–∑—É–µ–º –∫–æ–Ω—Å–µ—Ä–≤–∞—Ç–∏–≤–Ω—ã–π –ø–æ–¥—Ö–æ–¥
                        if actual_columns_count is None or actual_columns_count == 0:
                            actual_columns_count = min(max_cols_to_read, 20)  # –ú–∞–∫—Å–∏–º—É–º 20 –∫–æ–ª–æ–Ω–æ–∫
                        
                        # –¢–µ–ø–µ—Ä—å —Å–æ–∑–¥–∞—ë–º safe_visible_columns –Ω–∞ –æ—Å–Ω–æ–≤–µ —Ä–µ–∞–ª—å–Ω–æ–≥–æ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –∫–æ–ª–æ–Ω–æ–∫
                        safe_visible_columns = []
                        for col_idx in range(min(actual_columns_count, max_cols_to_read)):
                            if col_idx not in hidden_columns:
                                safe_visible_columns.append(col_idx)
                        
                        # –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞: —É–±–∏—Ä–∞–µ–º –ª—é–±—ã–µ –∏–Ω–¥–µ–∫—Å—ã >= actual_columns_count
                        safe_visible_columns = [col for col in safe_visible_columns if col < actual_columns_count]
                        
                        # –°–ø–æ—Å–æ–± 2: pandas —Å openpyxl engine (–ª—É—á—à–µ –¥–ª—è –≤—ã—á–∏—Å–ª–µ–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π —Ñ–æ—Ä–º—É–ª)
                        if content.empty:
                            try:
                                # Pandas –º–æ–∂–µ—Ç –ª—É—á—à–µ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å –≤—ã—á–∏—Å–ª–µ–Ω–Ω—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è —Ñ–æ—Ä–º—É–ª
                                content = pd.read_excel(
                                    BytesIO(xlsx_data),
                                    engine='openpyxl',
                                    nrows=max_rows_to_read,
                                    header=0,    # –ü–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ –∫–∞–∫ –∑–∞–≥–æ–ª–æ–≤–∫–∏
                                    usecols=safe_visible_columns,  # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –±–µ–∑–æ–ø–∞—Å–Ω—ã–µ –≤–∏–¥–∏–º—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                    na_values=['', ' ', 'nan', 'NaN', 'NaT', 'None', '#N/A', '#VALUE!', '#REF!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!']
                                )
                            except Exception as e2:
                                error_msgs.append(f"pandas+openpyxl: {str(e2)}")
                        
                        # –°–ø–æ—Å–æ–± 3: calamine engine (–±—ã—Å—Ç—Ä—ã–π –∏ –Ω–∞–¥–µ–∂–Ω—ã–π)
                        if content.empty:
                            try:
                                # –î–ª—è calamine –∏—Å–ø–æ–ª—å–∑—É–µ–º —Ç–æ–ª—å–∫–æ –±–µ–∑–æ–ø–∞—Å–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                content = pd.read_excel(
                                    BytesIO(xlsx_data),
                                    engine='calamine',
                                    nrows=max_rows_to_read,
                                    usecols=safe_visible_columns  # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –±–µ–∑–æ–ø–∞—Å–Ω—ã–µ –≤–∏–¥–∏–º—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                )
                            except Exception as e3:
                                error_msgs.append(f"calamine: {str(e3)}")
                        
                        # –°–ø–æ—Å–æ–± 4: xlrd engine (–¥–ª—è —Å—Ç–∞—Ä—ã—Ö —Ñ–∞–π–ª–æ–≤)
                        if content.empty and XLRD_AVAILABLE:
                            try:
                                content = pd.read_excel(
                                    BytesIO(xlsx_data),
                                    engine='xlrd',
                                    nrows=max_rows_to_read,
                                    usecols=safe_visible_columns  # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –±–µ–∑–æ–ø–∞—Å–Ω—ã–µ –≤–∏–¥–∏–º—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                )
                            except Exception as e4:
                                error_msgs.append(f"xlrd: {str(e4)}")
                        
                        # –°–ø–æ—Å–æ–± 5: –ü—Ä–æ—Å—Ç–æ–µ —á—Ç–µ–Ω–∏–µ –±–µ–∑ pandas (–ø–æ—Å–ª–µ–¥–Ω—è—è –ø–æ–ø—ã—Ç–∫–∞)
                        if content.empty and OPENPYXL_AVAILABLE:
                            try:
                                from openpyxl import load_workbook
                                
                                wb = load_workbook(BytesIO(xlsx_data), data_only=True, read_only=True)
                                ws = wb.active
                                
                                # –ß–∏—Ç–∞–µ–º –∫–∞–∫ –ø—Ä–æ—Å—Ç—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è —Å –ø—Ä–∞–≤–∏–ª—å–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–æ–π —Ñ–æ—Ä–º—É–ª
                                simple_data = []
                                if ws:  # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ worksheet –Ω–µ None
                                    for row in ws.iter_rows(values_only=True, max_row=max_rows_to_read, max_col=max_cols_to_read):
                                        # –í–∫–ª—é—á–∞–µ–º –≤—Å–µ —Å—Ç—Ä–æ–∫–∏, –≤–∫–ª—é—á–∞—è –ø—É—Å—Ç—ã–µ (–∫–∞–∫ –≤ Excel)
                                        processed_row = []
                                        for cell_idx, cell in enumerate(row):
                                            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–∫—Ä—ã—Ç—ã–µ –∫–æ–ª–æ–Ω–∫–∏
                                            if cell_idx in hidden_columns:
                                                continue
                                                
                                            if cell is None:
                                                processed_row.append("")
                                            else:
                                                # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ —Å—Ç—Ä–æ–∫—É –∏ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Ñ–æ—Ä–º—É–ª—ã
                                                cell_value = str(cell)
                                                # –£–±–∏—Ä–∞–µ–º –∞—Ä—Ç–µ—Ñ–∞–∫—Ç—ã pandas
                                                if cell_value in ['nan', 'NaN', 'NaT', 'None']:
                                                    processed_row.append("")
                                                else:
                                                    processed_row.append(cell_value)
                                        
                                        # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç—Ä–æ–∫–∏ —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –æ–Ω–∏ –Ω–µ –ø—É—Å—Ç—ã–µ –∏–ª–∏ —ç—Ç–æ –ø–µ—Ä–≤–∞—è —Å—Ç—Ä–æ–∫–∞ (–∑–∞–≥–æ–ª–æ–≤–∫–∏)
                                        if processed_row and (len(simple_data) == 0 or any(cell.strip() for cell in processed_row if cell)):
                                            simple_data.append(processed_row)
                                
                                wb.close()
                                
                                if simple_data:
                                    # –°–æ–∑–¥–∞–µ–º DataFrame —Å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–º–∏ –∑–∞–≥–æ–ª–æ–≤–∫–∞–º–∏
                                    if len(simple_data) > 0:
                                        headers = simple_data[0] if simple_data else []
                                        data_rows = simple_data[1:] if len(simple_data) > 1 else []
                                        content = pd.DataFrame(data_rows, columns=headers)
                                    else:
                                        content = pd.DataFrame()
                                    
                            except Exception as e5:
                                error_msgs.append(f"openpyxl –ø—Ä–æ—Å—Ç–æ–π: {str(e5)}")
                        
                        # –ï—Å–ª–∏ –≤—Å–µ —Å–ø–æ—Å–æ–±—ã –Ω–µ —Å—Ä–∞–±–æ—Ç–∞–ª–∏
                        if content.empty:
                            return (f"<p style='padding-top:8px;color:red;'>"
                                    f"<strong>{attachment.name}</strong><br/>"
                                    f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ—á–∏—Ç–∞—Ç—å xlsx —Ñ–∞–π–ª.<br/>"
                                    f"–û—à–∏–±–∫–∏: {' | '.join(error_msgs)}<br/><br/>"
                                    f"<strong>–í–æ–∑–º–æ–∂–Ω—ã–µ —Ä–µ—à–µ–Ω–∏—è:</strong><br/>"
                                    f"1. –§–∞–π–ª —Å–æ–¥–µ—Ä–∂–∏—Ç —Å–ª–æ–∂–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –∏–ª–∏ –ø–æ–≤—Ä–µ–∂–¥–µ–Ω<br/>"
                                    f"2. –ü–µ—Ä–µ—Å–æ—Ö—Ä–∞–Ω–∏—Ç–µ —Ñ–∞–π–ª –≤ Excel –∫–∞–∫ '–ö–Ω–∏–≥–∞ Excel (.xlsx)' –±–µ–∑ –º–∞–∫—Ä–æ—Å–æ–≤<br/>"
                                    f"3. –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–π—Ç–µ –≤ CSV —Ñ–æ—Ä–º–∞—Ç<br/>"
                                    f"4. –£–±–µ—Ä–∏—Ç–µ –≤—Å–µ —Å—Ç–∏–ª–∏, —Ñ–æ—Ä–º—É–ª—ã –∏ –∑–∞—â–∏—Ç—É –∏–∑ —Ñ–∞–π–ª–∞<br/>"
                                    f"5. –ü–æ–ø—Ä–æ–±—É–π—Ç–µ –æ—Ç–∫—Ä—ã—Ç—å —Ñ–∞–π–ª –≤ –¥—Ä—É–≥–æ–π –ø—Ä–æ–≥—Ä–∞–º–º–µ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Ü–µ–ª–æ—Å—Ç–Ω–æ—Å—Ç–∏</p>")
                        
                    elif doc_type == 'xls':
                        # –î–ª—è —Å—Ç–∞—Ä—ã—Ö .xls —Ñ–∞–π–ª–æ–≤ –∏—Å–ø–æ–ª—å–∑—É–µ–º xlrd
                        # –ê–Ω–∞–ª–∏–∑–∏—Ä—É–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É —Ñ–∞–π–ª–∞ –¥–ª—è –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏–∏ (–º–æ–∂–µ—Ç –Ω–µ —Ä–∞–±–æ—Ç–∞—Ç—å –¥–ª—è .xls)
                        try:
                            hidden_columns, max_data_column, max_data_row = self.get_hidden_columns_and_data_range(xlsx_data)
                            max_cols_to_read = min(max_data_column or 30, 50)  # –ú–∞–∫—Å–∏–º—É–º 50 –∫–æ–ª–æ–Ω–æ–∫
                            max_rows_to_read = min(max_data_row or 5000, 5000)  # –ú–∞–∫—Å–∏–º—É–º 5000 —Å—Ç—Ä–æ–∫
                            
                            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è —á—Ç–µ–Ω–∏—è (–∏—Å–∫–ª—é—á–∞–µ–º —Å–∫—Ä—ã—Ç—ã–µ)
                            visible_columns = []
                            for col_idx in range(max_cols_to_read):
                                if col_idx not in hidden_columns:
                                    visible_columns.append(col_idx)
                            
                            content = pd.read_excel(
                                BytesIO(xlsx_data), 
                                engine='xlrd',
                                nrows=max_rows_to_read,
                                usecols=visible_columns
                            )
                        except:
                            # –ï—Å–ª–∏ –∞–Ω–∞–ª–∏–∑ –Ω–µ —É–¥–∞–ª—Å—è, —á–∏—Ç–∞–µ–º –∫–∞–∫ –æ–±—ã—á–Ω–æ
                            content = pd.read_excel(
                                BytesIO(xlsx_data), 
                                engine='xlrd',
                                nrows=5000
                            )
                    elif doc_type == 'docx':
                        doc = DocxDocument(io.BytesIO(xlsx_data))
                        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                        if not paragraphs:
                            return ("<p style='padding-top:8px;color:orange;'>"
                                    "–î–æ–∫—É–º–µ–Ω—Ç –ø—É—Å—Ç –∏–ª–∏ –Ω–µ —Å–æ–¥–µ—Ä–∂–∏—Ç —Ç–µ–∫—Å—Ç–∞</p>")
                        return paragraphs
                    elif doc_type == 'pdf':
                        # –î–ª—è PDF –≤–æ–∑–≤—Ä–∞—â–∞–µ–º URL –¥–ª—è iframe –ø—Ä–æ—Å–º–æ—Ç—Ä–∞
                        return {
                            'type': 'pdf',
                            'url': f'/web/content/{attach_id}?download=false',
                            'attachment_id': attach_id,
                            'filename': attachment.name or 'document.pdf'
                        }
                    
                    # –î–ª—è Excel —Ñ–∞–π–ª–æ–≤ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º HTML —Ç–∞–±–ª–∏—Ü—É
                    if doc_type in ['xlsx', 'xls'] and content is not None:
                        if content.empty:
                            return ("<p style='padding-top:8px;color:orange;'>"
                                    "–§–∞–π–ª –Ω–µ —Å–æ–¥–µ—Ä–∂–∏—Ç –¥–∞–Ω–Ω—ã—Ö</p>")
                        
                        # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –≤—Å–µ –¥–∞–Ω–Ω—ã–µ –≤ –ø—Ä–µ–≤—å—é (–±–µ–∑ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π)
                        preview_note = ""
                        # –£–±–∏—Ä–∞–µ–º –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏—è –Ω–∞ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Å—Ç—Ä–æ–∫ –∏ –∫–æ–ª–æ–Ω–æ–∫
                        # if len(content) > 100:
                        #     content = content.head(100)
                        # 
                        # if len(content.columns) > 20:
                        #     content = content.iloc[:, :20]
                        
                        # –û—á–∏—â–∞–µ–º DataFrame –æ—Ç –Ω–µ–∂–µ–ª–∞—Ç–µ–ª—å–Ω—ã—Ö —ç–ª–µ–º–µ–Ω—Ç–æ–≤ —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ content –Ω–∞–π–¥–µ–Ω
                        if content is not None:
                            # 1. –ó–∞–º–µ–Ω—è–µ–º –Ω–∞–∑–≤–∞–Ω–∏—è –∫–æ–ª–æ–Ω–æ–∫ "Unnamed: 0", "Unnamed: 1" –Ω–∞ –ø—É—Å—Ç—ã–µ —Å—Ç—Ä–æ–∫–∏
                            new_columns = []
                            for col in content.columns:
                                if str(col).startswith('Unnamed:'):
                                    new_columns.append('')
                                else:
                                    new_columns.append(col)
                            content.columns = new_columns
                            
                            # 2. –ó–∞–º–µ–Ω—è–µ–º NaT, NaN, null –Ω–∞ –ø—É—Å—Ç—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è
                            content = content.fillna('')
                            content = content.replace('NaT', '')
                            content = content.replace('NaN', '')
                            content = content.replace('nan', '')
                        
                        html_table = content.to_html(index=False, escape=False)
                        return preview_note + html_table
                        
                except Exception as e:
                    return (f"<p style='padding-top:8px;color:red;'>"
                            f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ —Ñ–∞–π–ª–∞ {doc_type}: {str(e)}<br/><br/>"
                            f"<strong>–†–µ–∫–æ–º–µ–Ω–¥–∞—Ü–∏–∏:</strong><br/>"
                            f"1. –û—Ç–∫—Ä–æ–π—Ç–µ —Ñ–∞–π–ª –≤ Excel –∏ –ø–µ—Ä–µ—Å–æ—Ö—Ä–∞–Ω–∏—Ç–µ –±–µ–∑ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è<br/>"
                            f"2. –≠–∫—Å–ø–æ—Ä—Ç–∏—Ä—É–π—Ç–µ –≤ CSV –¥–ª—è –≥–∞—Ä–∞–Ω—Ç–∏—Ä–æ–≤–∞–Ω–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏<br/>"
                            f"3. –£–±–µ–¥–∏—Ç–µ—Å—å, —á—Ç–æ —Ñ–∞–π–ª –Ω–µ –∑–∞—â–∏—â–µ–Ω –ø–∞—Ä–æ–ª–µ–º</p>")
            
        except Exception as e:
            return (f"<p style='padding-top:8px;color:red;'>"
                    f"–û–±—â–∞—è –æ—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞: {str(e)}</p>")
        
        return ("<p style='padding-top:8px;color:red;'>"
                "–§–æ—Ä–º–∞—Ç —Ñ–∞–π–ª–∞ –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç—Å—è</p>") 